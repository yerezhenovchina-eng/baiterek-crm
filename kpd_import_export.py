# kpd_import_export.py — Импорт/экспорт КПД из Excel шаблонов
import openpyxl
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from datetime import datetime
import io

class KPDImporter:
    """Парсинг Excel файлов КПД по шаблонам"""
    
    @staticmethod
    def parse_personal_kpd(file_content, username):
        """
        Парсит личный КПД сотрудника из Excel
        Возвращает список (quarter, показатели)
        """
        wb = openpyxl.load_workbook(io.BytesIO(file_content))
        results = {}
        
        quarter_map = {
            '2кв КАРТА КПД': 'II',
            '3кв КАРТА КПД': 'III',
            '4кв КАРТА КПД': 'IV'
        }
        
        for sheet_name, quarter in quarter_map.items():
            if sheet_name not in wb.sheetnames:
                continue
                
            ws = wb[sheet_name]
            kpd_data = {
                'quarter': quarter,
                'employee': username,
                'indicators': []
            }
            
            # Читаем показатели (начиная со строки 8)
            for row in ws.iter_rows(min_row=8, max_row=ws.max_row, values_only=False):
                # Колонка A - номер показателя
                num_cell = row[0]
                if not num_cell.value or not isinstance(num_cell.value, int):
                    continue
                
                indicator = {
                    'num': num_cell.value,
                    'name': row[1].value or '',  # B - КПД
                    'unit': row[3].value or '',  # D - Ед.изм.
                    'formula': row[4].value or '',  # E - Формула
                    'formula_desc': row[5].value or '',  # F - Описание
                    'data_source': row[6].value or '',  # G - Источник
                    'weight': row[7].value or 0,  # H - Вес
                    'threshold': row[9].value or '',  # J - Порог
                    'goal': row[10].value or '',  # K - Цель
                    'challenge': None,  # У ДКВ нет этого поля
                    'fact': None  # Заполняется отдельно
                }
                
                kpd_data['indicators'].append(indicator)
            
            if kpd_data['indicators']:
                results[quarter] = kpd_data
        
        return results
    
    @staticmethod
    def parse_dkv_kpd(file_content):
        """
        Парсит КПД ДКВ (глобальный) из Excel
        Возвращает список (quarter, показатели)
        """
        wb = openpyxl.load_workbook(io.BytesIO(file_content))
        results = {}
        
        quarter_map = {
            '2кв КАРТА КПД': 'II',
            '3кв КАРТА КПД': 'III',
            '4кв КАРТА КПД': 'IV'
        }
        
        for sheet_name, quarter in quarter_map.items():
            if sheet_name not in wb.sheetnames:
                continue
                
            ws = wb[sheet_name]
            kpd_data = {
                'quarter': quarter,
                'employee': 'map',  # КПД ДКВ хранится как 'map'
                'indicators': []
            }
            
            # Читаем показатели (начиная со строки 8)
            for row in ws.iter_rows(min_row=8, max_row=ws.max_row, values_only=False):
                num_cell = row[0]
                if not num_cell.value or not isinstance(num_cell.value, int):
                    continue
                
                indicator = {
                    'num': num_cell.value,
                    'name': row[1].value or '',
                    'unit': row[3].value or '',
                    'formula': row[4].value or '',
                    'formula_desc': row[5].value or '',
                    'data_source': row[6].value or '',
                    'weight': row[7].value or 0,
                    'threshold': row[9].value or '',
                    'goal': row[10].value or '',
                    'challenge': None,
                    'fact': None
                }
                
                kpd_data['indicators'].append(indicator)
            
            if kpd_data['indicators']:
                results[quarter] = kpd_data
        
        return results


class KPDExporter:
    """Экспорт КПД в Excel по шаблонам"""
    
    @staticmethod
    def create_personal_kpd_excel(employee_name, employee_position, kpd_data_by_quarter):
        """
        Создаёт Excel файл личного КПД
        kpd_data_by_quarter = {
            'II': [indicators],
            'III': [indicators],
            'IV': [indicators]
        }
        """
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        
        quarter_names = {
            'II': '2кв КАРТА КПД',
            'III': '3кв КАРТА КПД',
            'IV': '4кв КАРТА КПД'
        }
        
        for quarter, sheet_name in quarter_names.items():
            if quarter not in kpd_data_by_quarter:
                continue
            
            ws = wb.create_sheet(sheet_name)
            indicators = kpd_data_by_quarter[quarter]
            
            # Заголовок
            ws['A1'] = f'КАРТА КПД НА {quarter} КВАРТАЛ 2026 ГОДА'
            ws['A2'] = 'ФИО'
            ws['D2'] = employee_name
            ws['A3'] = 'Должность'
            ws['D3'] = employee_position
            ws['A4'] = 'Структурное подразделение'
            ws['D4'] = 'Департамент корпоративного взаимодействия'
            ws['A5'] = 'РАЗДЕЛ I'
            ws['D5'] = 'КПД СТРУКТУРНОГО ПОДРАЗДЕЛЕНИЯ (ВЕС - 40%)'
            
            # Заголовки таблицы
            headers = ['№', 'КПД (не менее 2, не более 5)', None, 'Ед.изм.', 
                      'Формула расчета', 'Описание формулы', 'Источник данных', 
                      'Вес (%)', 'Плановая результативность', None, None, 
                      'Факт', 'Результативность', 'Обучение и профессиональное развитие', 
                      'Фактически реализованное обучение']
            
            for col, header in enumerate(headers, 1):
                if header:
                    ws.cell(row=6, column=col, value=header)
            
            # Подзаголовки для плановой результативности
            ws.cell(row=7, column=9, value='Порог')
            ws.cell(row=7, column=10, value='Цель')
            ws.cell(row=7, column=11, value='Вызов')
            
            # Показатели
            for i, ind in enumerate(indicators, 8):
                ws.cell(row=i, column=1, value=ind.get('num'))
                ws.cell(row=i, column=2, value=ind.get('name'))
                ws.cell(row=i, column=4, value=ind.get('unit'))
                ws.cell(row=i, column=5, value=ind.get('formula'))
                ws.cell(row=i, column=6, value=ind.get('formula_desc'))
                ws.cell(row=i, column=7, value=ind.get('data_source'))
                ws.cell(row=i, column=8, value=ind.get('weight'))
                ws.cell(row=i, column=9, value=ind.get('threshold'))
                ws.cell(row=i, column=10, value=ind.get('goal'))
                ws.cell(row=i, column=11, value=ind.get('challenge'))
                ws.cell(row=i, column=12, value=ind.get('fact'))
        
        return wb
    
    @staticmethod
    def create_dkv_kpd_excel(kpd_data_by_quarter):
        """
        Создаёт Excel файл КПД ДКВ
        """
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        
        quarter_names = {
            'II': '2кв КАРТА КПД',
            'III': '3кв КАРТА КПД',
            'IV': '4кв КАРТА КПД'
        }
        
        for quarter, sheet_name in quarter_names.items():
            if quarter not in kpd_data_by_quarter:
                continue
            
            ws = wb.create_sheet(sheet_name)
            indicators = kpd_data_by_quarter[quarter]
            
            # Заголовок
            ws['A1'] = f'КАРТА КПД НА {quarter} КВАРТАЛ 2026 ГОДА'
            ws['A2'] = 'ФИО'
            ws['A3'] = 'Должность'
            ws['A4'] = 'Структурное подразделение'
            ws['D4'] = 'Департамент корпоративного взаимодействия'
            ws['A5'] = 'РАЗДЕЛ I'
            ws['D5'] = 'КПД СТРУКТУРНОГО ПОДРАЗДЕЛЕНИЯ (ВЕС - 40%)'
            
            # Заголовки таблицы
            headers = ['№', 'КПД (не менее 2, не более 5)', None, 'Ед.изм.', 
                      'Формула расчета', 'Описание формулы', 'Источник данных', 
                      'Вес (%)', 'Плановая результативность', None, None, 
                      'Факт', 'Результативность', 'Обучение и профессиональное развитие', 
                      'Фактически реализованное обучение']
            
            for col, header in enumerate(headers, 1):
                if header:
                    ws.cell(row=6, column=col, value=header)
            
            # Подзаголовки
            ws.cell(row=7, column=9, value='Порог')
            ws.cell(row=7, column=10, value='Цель')
            
            # Показатели
            for i, ind in enumerate(indicators, 8):
                ws.cell(row=i, column=1, value=ind.get('num'))
                ws.cell(row=i, column=2, value=ind.get('name'))
                ws.cell(row=i, column=4, value=ind.get('unit'))
                ws.cell(row=i, column=5, value=ind.get('formula'))
                ws.cell(row=i, column=6, value=ind.get('formula_desc'))
                ws.cell(row=i, column=7, value=ind.get('data_source'))
                ws.cell(row=i, column=8, value=ind.get('weight'))
                ws.cell(row=i, column=9, value=ind.get('threshold'))
                ws.cell(row=i, column=10, value=ind.get('goal'))
                ws.cell(row=i, column=12, value=ind.get('fact'))
            
            # Итого строка
            total_row = 8 + len(indicators)
            ws.cell(row=total_row, column=2, value='Итого')
            ws.cell(row=total_row, column=8, value=f'=SUM(H8:H{total_row-1})')
            ws.cell(row=total_row, column=9, value=f'=SUM(I8:I{total_row-1})')
        
        return wb
